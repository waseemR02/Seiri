import openpyxl
from seiri.transform import Transform


def test_csv_to_xlsx(tmp_path):
    # Create a temporary output file path
    output_file = tmp_path / "output.xlsx"

    # Create a temporary log file path
    log_file = tmp_path / "seiri-error.log"

    # Create an instance of Transform
    transformer = Transform(log=str(log_file))

    # Call the csv_to_xlsx method
    transformer.csv_to_xlsx("tests/data/Sample.csv", str(output_file))

    # Check if the output file exists
    assert output_file.exists()

    # Open the output file using openpyxl
    wb = openpyxl.load_workbook(output_file)

    # Load the sample file
    sample_file = "tests/data/Sample.xlsx"
    sample_wb = openpyxl.load_workbook(sample_file)

    # Check if the workbook contains the expected sheets
    assert set(sample_wb.sheetnames) == set(wb.sheetnames)

    # Check if the sheets contain the expected data
    for sheet_name in sample_wb.sheetnames:
        sample_sheet = sample_wb[sheet_name]
        sheet = wb[sheet_name]
        assert sample_sheet.max_row == sheet.max_row
        assert sample_sheet.max_column == sheet.max_column
        for row in range(1, sample_sheet.max_row + 1):
            for col in range(1, sample_sheet.max_column + 1):
                assert (
                    sample_sheet.cell(row=row, column=col).value
                    == sheet.cell(row=row, column=col).value
                )

    # Close the sample workbook
    sample_wb.close()

    # Check if the log file is created
    assert log_file.exists()

    # Close the workbook
    wb.close()
