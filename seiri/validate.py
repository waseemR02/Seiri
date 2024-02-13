import openpyxl
from loguru import logger
import sys
import os


class Validate:
    """
    Validate checks given xlsx against given rules
    """

    def __init__(self, log: str = "seiri-error.log") -> None:
        # Initialize logger with default settings
        self.logger = logger
        self.logger.remove()
        self.logger.add(
            sink=log,
            level="ERROR",
            format="<white>{time:MMMM D, YYYY > HH:mm:ss}</white> | <level>{level: <8}</level> | <level>Validate</level> | <level>{message}</level>",
        )
        self.logger.add(
            sink=sys.stdout,
            format="<white>{time:MMMM D, YYYY > HH:mm:ss}</white> | <level>{level: <8}</level> | <level>Validate</level> |  <level>{message}</level>",
        )

    def validate(self, in_xlsx: str, against_xlsx: str) -> bool:
        """
        Validate checks given xlsx against given rules
        """
        # Return True if all guards pass
        ret = True

        if not os.path.exists(in_xlsx):
            self.logger.error(f"Could not find {in_xlsx}")
            ret = False

        if not os.path.exists(against_xlsx):
            self.logger.error(f"Could not find {against_xlsx}")
            ret = False

        self.logger.success(f"Found {in_xlsx} and {against_xlsx}")
        self.logger.info(f"Validating {in_xlsx} against {against_xlsx}")

        in_wb = openpyxl.load_workbook(in_xlsx)
        self.logger.success(f"Loaded {in_xlsx}")

        against_wb = openpyxl.load_workbook(against_xlsx)
        self.logger.success(f"Loaded {against_xlsx}")

        # 1. Check if both excel book row count is same in the en sheet
        self.logger.info(f"Checking row count against {against_xlsx} in en sheet")
        if in_wb["en"].max_row != against_wb["en"].max_row:
            self.logger.error("Row count mismatch in en sheet")
            ret = False
        else:
            self.logger.success("Row count match in en sheet")

        # 1. Check if key and value order is same in the en sheet for both excel sheet "en"
        self.logger.info(
            f"Checking key and value order against {against_xlsx} in en sheet"
        )
        for row in range(1, in_wb["en"].max_row + 1):
            if (
                in_wb["en"].cell(row=row, column=1).value
                != against_wb["en"].cell(row=row, column=1).value
            ):
                self.logger.error(
                    f"Key mismatch in en sheet at row {row}. Expected: {against_wb['en'].cell(row=row, column=1).value}, Actual: {in_wb['en'].cell(row=row, column=1).value}"
                )
                ret = False
            if (
                in_wb["en"].cell(row=row, column=2).value
                != against_wb["en"].cell(row=row, column=2).value
            ):
                self.logger.error(
                    f"Value mismatch in en sheet at row {row}. Expected: {against_wb['en'].cell(row=row, column=2).value}, Actual: {in_wb['en'].cell(row=row, column=2).value}"
                )
                ret = False

        self.logger.success("Key and Value order match in en sheet")
        self.logger.success(f"Validation successful against {against_xlsx}")

        # Now onto checks only in the in_xlsx
        # 2. Check if all sheets in the in_xlsx have same row count
        self.logger.info("Checking row count in all sheets")
        for sheet in in_wb.sheetnames:
            if in_wb[sheet].max_row != in_wb["en"].max_row:
                self.logger.error(f"Row count mismatch in {sheet} sheet")
                ret = False
            else:
                self.logger.success(f"Row count match in {sheet} sheet")

        # 3,4. The “en” sheet key should be available and match in all other sheets
        self.logger.info("Checking if 'en' key is available in all sheets")
        for row in range(1, in_wb["en"].max_row + 1):
            key = in_wb["en"].cell(row=row, column=1).value
            # Check if this key is in all other sheets
            for sheet in in_wb.sheetnames:
                if sheet == "en":
                    continue

                # handle when cell is empty
                if in_wb[sheet].cell(row=row, column=1).value is None:
                    self.logger.error(f"Empty cell in {sheet} sheet at row {row}")
                    continue

                if key not in in_wb[sheet].cell(row=row, column=1).value:
                    self.logger.error(
                        f"Mismatched value in {sheet} sheet at row {row}. Expected: {key}, Actual: {in_wb[sheet].cell(row=row, column=1).value}"
                    )
                    ret = False
        self.logger.success(
            "Key found in all sheets and the order is same across all sheets"
        )

        # 5. Check if the value text in all other sheets should not be empty
        self.logger.info("Checking if value is not empty in all sheets")
        for sheet in in_wb.sheetnames:
            if sheet == "en":
                continue
            for row in range(1, in_wb[sheet].max_row + 1):
                if in_wb[sheet].cell(row=row, column=2).value == "":
                    self.logger.error(
                        f"Empty value found in {sheet} sheet at row {row}"
                    )
                    ret = False
        self.logger.success("Value Column is not empty in all sheets - 'en'")

        # 6. Check if the text in value column in all sheets is unique
        self.logger.info("Checking if value is unique in all sheets")
        for sheet in in_wb.sheetnames:
            unique_values = set()
            for row in range(1, in_wb[sheet].max_row + 1):
                value = in_wb[sheet].cell(row=row, column=2).value
                if value in unique_values:
                    self.logger.error(
                        f"Duplicate value '{value}' found in {sheet} sheet"
                    )
                    ret = False
                unique_values.add(value)
        self.logger.success("Value Column is unique in all sheets")

        # 8. Check if the string length of value column in all sheets is less than the string length if value column in en sheet
        self.logger.info("Checking if value is less than en value in all sheets")
        for sheet in in_wb.sheetnames:
            if sheet == "en":
                continue
            for row in range(1, in_wb[sheet].max_row + 1):
                # When the cell is empty
                if in_wb[sheet].cell(row=row, column=2).value is None:
                    continue

                if len(in_wb[sheet].cell(row=row, column=2).value) > len(
                    in_wb["en"].cell(row=row, column=2).value
                ):
                    value = in_wb[sheet].cell(row=row, column=2).value
                    en_value = in_wb["en"].cell(row=row, column=2).value
                    self.logger.error(
                        f"Length({value})={len(value)} in '{sheet}' sheet at row {row} is greater than length({en_value})={len(en_value)} in 'en' sheet"
                    )
                    ret = False
        self.logger.success("Value Column is less than en value in all sheets")

        self.logger.success(f"Validation successful for {in_xlsx}")

        return ret


if __name__ == "__main__":
    import argparse

    ap = argparse.ArgumentParser()
    # get the first arg for in_xlsx without any flag
    ap.add_argument(
        "in_xlsx",
        type=str,
        default="tests/data/Delivered_correct.xlsx",
        help="path to xlsx to validate",
    )
    ap.add_argument(
        "--against",
        type=str,
        default="tests/data/Sample.xlsx",
        help="path to xlsx to validate against",
    )

    ap.add_argument(
        "--log",
        type=str,
        default="seiri-error.log",
        help="path to log file",
    )

    args = vars(ap.parse_args())

    validate = Validate(args["log"])
    validate.validate(args["in_xlsx"], args["against"])
